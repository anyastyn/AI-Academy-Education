import os
import glob
import requests
from dotenv import load_dotenv

from docx import Document
from openpyxl import load_workbook

import httpx
from openai import OpenAI

import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SERVICE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")

if not all([SUPABASE_URL, SERVICE_KEY, OPENAI_API_KEY]):
    raise SystemExit("Missing env vars: SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY, OPENAI_API_KEY")

# Corporate SSL workaround (OpenAI uses httpx)
http_client = httpx.Client(verify=False, timeout=60.0)
client = OpenAI(api_key=OPENAI_API_KEY, http_client=http_client)

headers = {
    "apikey": SERVICE_KEY,
    "Authorization": f"Bearer {SERVICE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation",
}

RAG_FOLDER = "RAG Data"


# ---------------------------
# Supabase REST helpers
# ---------------------------
def supabase_post(url: str, payload):
    r = requests.post(url, headers=headers, json=payload, timeout=60, verify=False)
    r.raise_for_status()
    return r.json()

def supabase_get(url: str, params=None):
    r = requests.get(url, headers=headers, params=params, timeout=60, verify=False)
    r.raise_for_status()
    return r.json()

def delete_rows(table: str, where: dict):
    """
    where is a dict of query params like:
      {"document_id": "eq.<uuid>"}
      {"id": "eq.<uuid>"}
    """
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    r = requests.delete(url, headers=headers, params=where, timeout=60, verify=False)
    r.raise_for_status()

def insert_row(table: str, row: dict) -> dict:
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    return supabase_post(url, [row])[0]

def find_document_by_source(source_path: str):
    url = f"{SUPABASE_URL}/rest/v1/knowledge_documents"
    rows = supabase_get(url, params={"source": f"eq.{source_path}"})
    return rows[0] if rows else None


# ---------------------------
# Chunking + Embeddings
# ---------------------------
def chunk_text(text: str, chunk_size: int = 900, overlap: int = 120):
    text = (text or "").strip()
    if not text:
        return []
    chunks = []
    start = 0
    while start < len(text):
        end = min(len(text), start + chunk_size)
        chunks.append(text[start:end])
        if end == len(text):
            break
        start = max(0, end - overlap)
    return chunks

def embed_texts(texts):
    emb = client.embeddings.create(model="text-embedding-3-small", input=texts)
    return [d.embedding for d in emb.data]

def to_pgvector(vec):
    return "[" + ",".join(str(x) for x in vec) + "]"


# ---------------------------
# File readers
# ---------------------------
def read_docx_as_text(path: str) -> str:
    doc = Document(path)
    parts = []
    for p in doc.paragraphs:
        t = (p.text or "").strip()
        if t:
            parts.append(t)
    return "\n".join(parts)

def read_text_file(path: str) -> str:
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()

def read_xlsx_as_row_chunks(path: str):
    """
    IMPORTANT:
    Excel is chunked by ROW so retrieval works well for specific connector names like "SharePoint".
    """
    wb = load_workbook(path, data_only=True)
    row_chunks = []

    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        header = [str(c).strip() for c in rows[0] if c is not None]
        header = header if header else []

        for r in rows[1:]:
            cells = [("" if c is None else str(c).strip()) for c in r]
            if not any(cells):
                continue

            if header and len(header) <= len(cells):
                pairs = []
                for i, h in enumerate(header):
                    v = cells[i] if i < len(cells) else ""
                    if h and v:
                        pairs.append(f"{h}: {v}")
                chunk = " | ".join(pairs) if pairs else " | ".join([c for c in cells if c])
            else:
                chunk = " | ".join([c for c in cells if c])

            chunk = f"[Sheet: {sheet.title}] {chunk}"
            row_chunks.append(chunk)

    return row_chunks

def load_file_chunks(path: str):
    ext = os.path.splitext(path)[1].lower()

    if ext == ".docx":
        text = read_docx_as_text(path)
        return chunk_text(text)

    if ext == ".xlsx":
        return read_xlsx_as_row_chunks(path)

    if ext in [".txt", ".md"]:
        text = read_text_file(path)
        return chunk_text(text)

    raise ValueError(f"Unsupported file type: {ext}")


# ---------------------------
# Main ingestion
# ---------------------------
def main():
    if not os.path.isdir(RAG_FOLDER):
        raise SystemExit(f"Folder not found: '{RAG_FOLDER}' at repo root.")

    all_paths = (
        glob.glob(os.path.join(RAG_FOLDER, "*.docx")) +
        glob.glob(os.path.join(RAG_FOLDER, "*.xlsx")) +
        glob.glob(os.path.join(RAG_FOLDER, "*.txt")) +
        glob.glob(os.path.join(RAG_FOLDER, "*.md"))
    )

    # Skip Office temp/lock files like "~$Approved Connectors.xlsx"
    paths = sorted([
        p for p in all_paths
        if not os.path.basename(p).startswith("~$")
    ])

    if not paths:
        raise SystemExit(f"No supported files found in '{RAG_FOLDER}' (.docx/.xlsx/.txt/.md)")

    print(f"=== Ingesting knowledge files from: {RAG_FOLDER}/ ===")
    print(f"Found {len(paths)} files.")

    for path in paths:
        filename = os.path.basename(path)
        print(f"\n--- {filename} ---")

        # Overwrite behavior per file: delete old + re-ingest
        existing = find_document_by_source(path)
        if existing:
            print("Found existing document in Supabase → cleaning old chunks...")
            delete_rows("knowledge_chunks", {"document_id": f"eq.{existing['id']}"})
            delete_rows("knowledge_documents", {"id": f"eq.{existing['id']}"})

        # Insert document row
        doc_row = insert_row("knowledge_documents", {
            "title": filename,
            "source": path,
            "metadata": {"folder": RAG_FOLDER, "type": os.path.splitext(filename)[1].lower()}
        })
        doc_id = doc_row["id"]

        # Load chunks
        chunks = load_file_chunks(path)
        chunks = [c.strip() for c in chunks if (c or "").strip()]
        if not chunks:
            print("⚠️ No usable content found. Skipping.")
            continue

        # Embed in batches
        all_embeddings = []
        batch_size = 64
        for i in range(0, len(chunks), batch_size):
            batch = chunks[i:i + batch_size]
            all_embeddings.extend(embed_texts(batch))

        # Insert chunks
        for i, (chunk, emb) in enumerate(zip(chunks, all_embeddings)):
            insert_row("knowledge_chunks", {
                "document_id": doc_id,
                "chunk_index": i,
                "content": chunk,
                "embedding": to_pgvector(emb),
                "metadata": {"filename": filename}
            })

        print(f"✅ Ingested {filename}: {len(chunks)} chunks")

    print("\n🎉 Done. Knowledge base is ready for RAG.")

if __name__ == "__main__":
    main()
